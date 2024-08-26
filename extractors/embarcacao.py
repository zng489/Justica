import argparse
import datetime
import re
from unicodedata import normalize

import rows
import xlrd
from openpyxl import load_workbook

REGEXP_SPACES = re.compile(r"\s+")


def unaccent(value):
    return normalize("NFKD", value).encode("ascii", errors="ignore").decode("ascii")


class DocumentField(rows.fields.TextField):
    @classmethod
    def deserialize(cls, value):
        value = str(value or "").strip()
        if not value:
            return None
        return value.replace("  ", "")


class PtBrDateField(rows.fields.DateField):
    @classmethod
    def deserialize(cls, value):
        value = str(value or "").replace("00:00:00", "").strip()
        if not value:
            return None
        elif value.count("/") == 2:
            return datetime.datetime.strptime(value, "%d/%m/%Y").date()
        elif value.count("/") == 1:  # Weird case: '2810/2014'
            return datetime.datetime.strptime(value, "%d%m/%Y").date()
        elif value.count("-") == 2:
            return datetime.datetime.strptime(value, "%Y-%m-%d").date()
        else:
            raise ValueError(f"Unknown date: {repr(value)}")


class FileConverter:
    STATUS_CHOICES = ("CANCELADO", "EXPIRADA")

    def __init__(self, filename):
        self.filename = filename

    @property
    def _boundaries_xls(self):
        wb = xlrd.open_workbook(self.filename)
        sheet = wb.sheet_by_index(0)
        start_index = stop_index = None
        for index in range(sheet.nrows):
            row = sheet.row(index)
            row_text = " ".join(str(cell.value or "").strip() for cell in row).strip()
            if not row_text:
                continue
            elif start_index is None and row_text.startswith("Nº REB"):
                start_index = index
            elif start_index is not None and ("DADOS ATUALIZADOS" in row_text or "TOTAL TOTAL" in row_text):
                stop_index = index - 1
                break
        return start_index, stop_index

    @property
    def _boundaries_xlsx(self):
        wb = load_workbook(self.filename)
        sheet = wb[wb.sheetnames[0]]
        start_index = stop_index = None
        for index, row in enumerate(sheet.rows):
            row_text = " ".join(str(cell.value or "").strip() for cell in row).strip()
            if not row_text:
                continue
            elif start_index is None and row_text.startswith("Nº REB"):
                start_index = index
            elif start_index is not None and ("DADOS ATUALIZADOS" in row_text or "TOTAL TOTAL" in row_text):
                stop_index = index - 1
                break
        return start_index, stop_index

    @property
    def _boundaries(self):
        filename = self.filename.lower()
        if filename.endswith(".xls"):
            return self._boundaries_xls
        elif filename.endswith(".xlsx"):
            return self._boundaries_xlsx

    @property
    def data(self):
        start_index, stop_index = self._boundaries
        filename = self.filename.lower()
        if filename.endswith(".xls"):
            func = rows.import_from_xls
        elif filename.endswith(".xlsx"):
            func = rows.import_from_xlsx
        table = func(
            self.filename,
            start_row=start_index,
            end_row=stop_index,
            force_types={
                "data_registro": PtBrDateField,
                "rpm_tie_ait": DocumentField,
            },
        )
        for row in table:
            data_validade = str(row.v_a_l_i_d_a_d_e or "").strip().upper()
            status = None
            if data_validade in self.STATUS_CHOICES:
                status, data_validade = data_validade.title(), None
            else:
                data_validade = PtBrDateField.deserialize(data_validade)

            yield {
                "reb": row.no_reb,
                "nome": row.nome_da_embarcacao,
                "data_registro": row.data_registro,
                "data_validade": data_validade,
                "proprietario": unaccent(row.proprietario_afretador),
                "rpm_tie_ait": REGEXP_SPACES.sub(" ", row.no_rpm_tie_ait),
                "status": status,
            }

    def to_csv(self, csv_filename):
        writer = rows.utils.CsvLazyDictWriter(csv_filename)
        for row in self.data:
            writer.writerow(row)
        writer.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("input_filename")
    parser.add_argument("output_filename")
    args = parser.parse_args()

    converter = FileConverter(args.input_filename)
    converter.to_csv(args.output_filename)
